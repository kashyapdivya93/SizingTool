#---------------------------------------------------------------------------------
# Script name : ReadServerInputs.py
# Author      : Nitin Bhagat
# Date        : 16-Jan-2020
# Purpose     : Sizing Tool Automation
# Parameters  : The script should read all the inputs needed for DB<->Server Mappings
# Parent file : WrapperScript.py
#---------------------------------------------------------------------------------

import os
import sys
import re
import csv
import math
import openpyxl
import win32com.client
import time

stage_directory = sys.argv[1]
outFilesString = sys.argv[2]
environmentName = sys.argv[3]
IOGrowth = sys.argv[4]
workbookFile = sys.argv[5]
outFiles = list(outFilesString.split(" "))
row = [[]]

# ---------------------------------------------------------------------------------
# Round function for consistent rounding
# ---------------------------------------------------------------------------------
def normal_round(n):
    if n - math.floor(n) < 0.5:
        return math.floor(n)
    return math.ceil(n)

# ---------------------------------------------------------------------------------
# Function to assign 0 to a variable if empty
# ---------------------------------------------------------------------------------
def assignZero(n):
    if len(n)==0:
        return 0
    else:
        return n

# ---------------------------------------------------------------------------------
# Function to store required values to variable to append to workbook file
# ---------------------------------------------------------------------------------
def insertToCSV(databaseName, instances, server, IOGrowth, read_iops, write_iops, sga, pga):
    row.append([databaseName, instances, server, IOGrowth, read_iops, write_iops, sga, pga])


def getAllValues(f1):
    for line in f1:
        if re.search(r'^DB_NAME', line):
            databaseName = line
            databaseName = databaseName[8:]
            databaseName = databaseName.strip()
        if re.search(r'^INSTANCES', line):
            instances = line
            instances = instances[9:]
            instances = int(instances.strip())
        if re.search(r'^HOSTS', line):
            server = line
            server = server[5:]
            server = server.strip()
            server_list = server.split(',')
    print("\n\n---------------------------------------------------------------------------------")
    print("Database Name: %s" % databaseName)
    print("Instances: %d" % instances)
    print("Host: %s" % server)
    
    # ---------------------------------------------------------------------------------
    # Get BEGIN-MAIN-METRICS & END-MAIN-METRICS line number
    # ---------------------------------------------------------------------------------
    begin_main_metrics = 0
    end_main_metrics = 0
    iops_check_string = 0
    for num, line in enumerate(f1, 1):
        if '~~BEGIN-MAIN-METRICS~~' in line:
            begin_main_metrics = num + 3
            iops_check_string = num + 1 # Line where string exists
        if '~~END-MAIN-METRICS~~' in line:
            end_main_metrics = num-2
            break
    
    start_read_iops_max = f1[iops_check_string].find("read_iops") + 1 # Find the starting position of string "read_iops_max"
    end_read_iops_max = start_read_iops_max + len("read_iops") - 1 # Find the ending position of string "read_iops_max"
    
    start_write_iops_max = f1[iops_check_string].find("write_iops") + 1 # Find the starting position of string "write_iops_max"
    end_write_iops_max = start_write_iops_max + len("write_iops") - 1 # Find the ending position of string "write_iops_max"
    
    start_snap = f1[iops_check_string].find("snap") - 4 # Find the starting position of string "snap"
    end_snap = start_snap + 9 # Find the ending position of string "snap"

    print("File %s" % outFile)
    print("Fetching read_iops and write_iops ...")
    read_iops_list = {}
    write_iops_list = {}
    # all_read_iops = []
    # all_write_iops = []
    max_read_iops = 0
    max_write_iops = 0
    # Iterate through the above 2 lines
    for i in range(begin_main_metrics, end_main_metrics):
        # Gets the entire line
        line = f1[i]
        # Slices the line to include only read_iops and write_iops. Replace ',' with '.' (if file is corrupted)
        read_iops = line[start_read_iops_max:end_read_iops_max]
        read_iops = read_iops.replace(',', '.')
        write_iops = line[start_write_iops_max:end_write_iops_max]
        write_iops = write_iops.replace(',', '.')
        # Removes extra spaces 
        read_iops = read_iops.strip()
        write_iops = write_iops.strip()
        # Convert from string to float
        read_iops = float(read_iops)
        write_iops = float(write_iops)
        # Do above for snap, add to dictionary as key and give a initial value of 0
        snap = line[start_snap:end_snap]
        snap = snap.replace(',', '.')
        snap = snap.strip()
        snap = int(snap)
        if snap not in read_iops_list:
            read_iops_list[snap] = 0
        if snap not in write_iops_list:
            write_iops_list[snap] = 0
        
        # Add read_iops and write_iops to list for the particular snap ID
        read_iops_list[snap] += read_iops
        write_iops_list[snap] += write_iops

    print("Finding max of read_iops and write_iops grouped by snap ID ...")
    keymax_read_iops = max(read_iops_list.keys(), key=(lambda k: read_iops_list[k]))
    max_read_iops = read_iops_list[keymax_read_iops]
    keymax_write_iops = max(write_iops_list.keys(), key=(lambda k: write_iops_list[k]))
    max_write_iops = write_iops_list[keymax_write_iops]

    time.sleep(3)

    # ---------------------------------------------------------------------------------
    # Get BEGIN-MEMORY & END-MEMORY line number
    # ---------------------------------------------------------------------------------
    begin_memory = 0
    end_memory = 0
    check_string = 0
    for num, line in enumerate(f1, 1):
        if '~~BEGIN-MEMORY~~' in line:
            begin_memory = num + 3
            check_string = num + 1 # Line where string exists
        if '~~END-MEMORY~~' in line:
            end_memory = num-2
            break
    
    start_sga = f1[check_string].find("SGA") - 3 # Find the starting position of string "SGA"
    end_sga = start_sga + len("SGA") + 3 # Find the ending position of string "SGA"

    start_pga = f1[check_string].find("PGA") - 3 # Find the starting position of string "PGA"
    end_pga = start_pga + len("PGA") + 3 # Find the ending position of string "PGA"

    start_snap = f1[check_string].find("SNAP_ID") # Find the starting position of string "snap"
    end_snap = start_snap + 7 # Find the ending position of string "snap"

    print("Fetching SGA and PGA ...")
    sga_list = {}
    pga_list = {}
    max_sga = 0
    max_pga = 0
    # Iterate through the above 2 lines
    for i in range(begin_memory, end_memory):
        # Gets the entire line
        line = f1[i]
        # Slices the line to include only SGAs & PGAs. Replace ',' with '.' (if file is corrupted)
        sga = line[start_sga:end_sga]
        sga = sga.replace(',', '.')
        pga = line[start_pga:end_pga]
        pga = pga.replace(',', '.')
        # Removes extra spaces
        sga = sga.strip()
        pga = pga.strip()
        # Convert from string to float
        sga = float(assignZero(sga))
        pga = float(assignZero(pga))
        
        # Do above for snap, add to dictionary as key and give a initial value of 0
        snap = line[start_snap:end_snap]
        snap = snap.replace(',', '.')
        snap = snap.strip()
        snap = int(snap)
        if snap not in sga_list:
            sga_list[snap] = 0
        if snap not in pga_list:
            pga_list[snap] = 0
        
        # Add read_iops and write_iops to list for the particular snap ID
        sga_list[snap] += sga
        pga_list[snap] += pga

    print("Finding max of SGA and PGA grouped by snap ID ...")
    print("---------------------------------------------------------------------------------")
    keymax_sga = max(sga_list.keys(), key=(lambda k: sga_list[k]))
    max_sga = sga_list[keymax_sga]
    keymax_pga = max(pga_list.keys(), key=(lambda k: pga_list[k]))
    max_pga = pga_list[keymax_pga]

    time.sleep(3)

    # insert all values to CSV file
    for i in range(instances):
        insertToCSV(databaseName, instances, server_list[i], IOGrowth, normal_round(max_read_iops/instances), normal_round(max_write_iops/instances), normal_round(max_sga/instances), normal_round(max_pga/instances))

# ---------------------------------------------------------------------------------
# Function to validate the mappings
# ---------------------------------------------------------------------------------
def validateMappingSheet():
    try:
        print("Validating inputs.... Please click on the excel dialog box")
        xlApp = win32com.client.DispatchEx('Excel.Application')
        xlsPath = os.path.expanduser(
            workbookFile)
        wb = xlApp.Workbooks.Open(Filename=xlsPath)
        xlApp.Run('Validate')
        wb.Save()
        xlApp.Quit()
        print("Macro ran successfully!")
    except Exception as e:
        print(e)
        print("Error found while running the excel macro!")
        xlApp.Quit()

# ---------------------------------------------------------------------------------
# Main Function
# ---------------------------------------------------------------------------------
for outFile in outFiles:
    file = open(stage_directory + '\\' + outFile)
    f1 = file.readlines()
    getAllValues(f1)
    file.close()

# ---------------------------------------------------------------------------------
# Append to Workbook file
# ---------------------------------------------------------------------------------
wbk = openpyxl.load_workbook(
    filename=workbookFile, read_only=False, keep_vba=True)
wks = wbk['DB<->Server Mappings']
print("\n\n---------------------------------------------------------------------------------")
print("Inserting values in Workbook File ...")
print("---------------------------------------------------------------------------------")
for r in range(1, len(row)):
    x = r + 2
    wks.cell(row=x, column=3).value = row[r][2]
    wks.cell(row=x, column=4).value = row[r][3]
    wks.cell(row=x, column=5).value = row[r][4]
    wks.cell(row=x, column=6).value = row[r][5]
    wks.cell(row=x, column=11).value = row[r][6]
    wks.cell(row=x, column=12).value = row[r][7]
wbk.save(workbookFile)
wbk.close

print("\n\n---------------------------------------------------------------------------------")
print("Validate mapping sheet DB<->Server Mappings")
validateMappingSheet()
print("---------------------------------------------------------------------------------")

print("\n\n---------------------------------------------------------------------------------")
print("Removing junk values under required columns ...")
print("---------------------------------------------------------------------------------")
wbk = openpyxl.load_workbook(filename=workbookFile, read_only=False, keep_vba=True)
wks = wbk['DB<->Server Mappings']

for r in range(1, len(row)):
    x = r + 2
    wks.cell(row=x, column=7).value = None
    wks.cell(row=x, column=8).value = None

wbk.save(workbookFile)
wbk.close

print("\n\n---------------------------------------------------------------------------------")
print("ReadServerInputs.py SCRIPT END!")
print("---------------------------------------------------------------------------------")

time.sleep(3)