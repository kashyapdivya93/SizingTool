#---------------------------------------------------------------------------------
# Script name : ReadDatabaseInputs.py
# Author      : Nitin Bhagat, Pranav Sriram
# Date        : 16-Jan-2020
# Purpose     : Sizing Tool Automation
# Parameters  : The script should read all the inputs needed for DB<->Server Mappings
# Parent file : WrapperScript.py
#---------------------------------------------------------------------------------

import openpyxl
import csv
import re
import sys
import subprocess
import win32com.client
import os

stage_directory = sys.argv[1]
outFilesString = sys.argv[2]
outFiles = list(outFilesString.split(" "))
environmentName = sys.argv[3]
plannedDataGrowth = sys.argv[4]
workbookFile = sys.argv[5]
row = [[]]

# ---------------------------------------------------------------------------------
# Function to store required values to variable to append to workbook file
# ---------------------------------------------------------------------------------
def insertToCSV(databaseName, instances, dbVersion, clustered, dbSize):
    row.append([environmentName, databaseName, instances, "ORACLE",
                dbVersion, clustered.upper(), dbSize.replace(',', '.'), plannedDataGrowth])


def getAllValues(out_file):
    for line in out_file:
        if re.search(r'^DB_NAME*', line):
            databaseName = line
            databaseName = databaseName[7:]
        if re.search(r'^INSTANCES*', line):
            instances = line
            instances = instances[9:]
        if re.search(r'^VERSION*', line):
            dbVersion = line
            dbVersion = dbVersion[7:]

    # ---------------------------------------------------------------------------------
    # Get BEGIN-SIZE-ON-DISK & END-SIZE-ON-DISK line number
    # ---------------------------------------------------------------------------------
    endMainMetrics, beginMainMetrics = 0, 0
    for num, line in enumerate(out_file, 1):
        if '~~BEGIN-SIZE-ON-DISK~~' in line:
            beginMainMetrics = num + 4
        if '~~END-SIZE-ON-DISK~~' in line:
            endMainMetrics = num - 3

    dbReadLine = out_file[endMainMetrics]
    dbSize = dbReadLine[11:]
    # clustered / non-clustered / shared-nothing
    clustered = "clustered" if int(instances) > 1 else "non-clustered"

    # ---------------------------------------------------------------------------------
    # Converting the version from 11.2.0.3 to 11g and respective values
    # ---------------------------------------------------------------------------------
    version = dbVersion.strip()
    version = version.rsplit('.', 1)[0]
    version = version.replace('.', '')
    version = int(version)
    if version >= 7000 and version < 7999:
        dbVersion = "7.3"
    elif version >= 8000 and version < 8999:
        dbVersion = "8i"
    elif version >= 9000 and version < 9999:
        dbVersion = "9i"
    elif version >= 10000 and version < 10999:
        dbVersion = "10g"
    elif version >= 11000 and version < 11999:
        dbVersion = "11g"
    elif version >= 12000 and version < 12999:
        dbVersion = "12c"
    elif version >= 18000 and version < 19999:
        dbVersion = "18c"
    else:
        dbVersion = "19c"

    insertToCSV(databaseName.strip(), instances.strip(), dbVersion,
                clustered.strip(), dbSize.strip())

# ---------------------------------------------------------------------------------
# Function to validate the mappings
# ---------------------------------------------------------------------------------
def validateMappingSheet():
    try:
        print("Validating inputs.... Please click on the excel dialog box")
        xlApp = win32com.client.DispatchEx('Excel.Application')
        xlsPath = os.path.expanduser(
            workbookFile)
        wb = xlApp.Workbooks.Open(Filename=xlsPath)
        xlApp.Run('Validate')
        wb.Save()
        xlApp.Quit()
        print("Macro ran successfully!")
    except Exception as e:
        print(e)
        print("Error found while running the excel macro!")
        xlApp.Quit()

# ---------------------------------------------------------------------------------
# Main Function
# ---------------------------------------------------------------------------------
for outFile in outFiles:
    file = open(stage_directory + '\\' + outFile)
    f1 = file.readlines()
    getAllValues(f1)
    file.close()

# ---------------------------------------------------------------------------------
# Append to Workbook file
# ---------------------------------------------------------------------------------
wbk = openpyxl.load_workbook(
    filename=workbookFile, read_only=False, keep_vba=True)
wks = wbk['Databases']
print("\n\n---------------------------------------------------------------------------------")
print("Inserting values in Workbook File ...")
print("---------------------------------------------------------------------------------")
for r in range(1, len(row)):
    x = r + 1
    wks.cell(row=x, column=1).value = row[r][0]
    wks.cell(row=x, column=2).value = row[r][1]
    wks.cell(row=x, column=3).value = row[r][2]
    wks.cell(row=x, column=4).value = row[r][3]
    wks.cell(row=x, column=5).value = row[r][4]
    wks.cell(row=x, column=6).value = row[r][5]
    wks.cell(row=x, column=7).value = row[r][6]
    wks.cell(row=x, column=15).value = row[r][7]
wbk.save(workbookFile)
wbk.close

wbk = openpyxl.load_workbook(
    filename=workbookFile, read_only=False, keep_vba=True)
wks = wbk['DB<->Server Mappings']
for r in range(1, len(row)):
    x = r + 1
    wks.cell(row=x, column=1).value = row[r][1]
wbk.save(workbookFile)
wbk.close

print("\n\n---------------------------------------------------------------------------------")
print("Validate mapping sheet DB<->Server Mappings")
validateMappingSheet()
print("---------------------------------------------------------------------------------")

print("\n\n---------------------------------------------------------------------------------")
print("ReadDatabaseInputs.py SCRIPT END!")
print("---------------------------------------------------------------------------------")