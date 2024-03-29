#---------------------------------------------------------------------------------
# Script name : WrapperScript.py
# Author      : Nitin Bhagat, Pranav Sriram
# Date        : 16-Jan-2020
# Purpose     : Sizing Tool Automation
# Parameters  : The script takes 0 parameters
# Usage       : python3 WrapperScript.py
#---------------------------------------------------------------------------------

import datetime
startTime = datetime.datetime.now()
import time
import subprocess
from subprocess import call
import sys
import re
import math
import csv
import openpyxl
import statistics
import os
import shutil
import win32com.client

# ---------------------------------------------------------------------------------
# Function to convert list to string to pass as a command line argument
# ---------------------------------------------------------------------------------
def convertToString(listToString):
    return ' '.join(listToString)

# ---------------------------------------------------------------------------------
# Round function for consistent rounding
# ---------------------------------------------------------------------------------
def normal_round(n):
    if n - math.floor(n) < 0.5:
        return math.floor(n)
    return math.ceil(n)

# ---------------------------------------------------------------------------------
# Function to assign 0 to a variable if empty
# ---------------------------------------------------------------------------------
def assignZero(n):
    if len(n)==0:
        return 0
    else:
        return n

# ---------------------------------------------------------------------------------
# Function to find workbook file
# ---------------------------------------------------------------------------------
def findWorkbookfile(name, location):
    for root, dirs, files in os.walk(location):
        if name in files:
            return os.path.join(root, name)

# ---------------------------------------------------------------------------------
# Function to calculate difference between startTime and endTime
# ---------------------------------------------------------------------------------
def date_diff_in_Minutes(dt2, dt1):
  timedelta = dt2 - dt1
  return timedelta.days * 24 * 60 + (timedelta.seconds/60)

os.system("cls")
print("---------------------------------------------------------------------------------")
print("Copying Main Workbook File to new location in newWorkbookFile directory")
print("Renaming copied workbook file")
print("---------------------------------------------------------------------------------")
location = os.getcwd()
name = "OfflineCustomerWorkbook-v200330081105.xlsm"
workbookFile = findWorkbookfile(name, location)
timestampStr = datetime.datetime.now().strftime("%d-%b-%Y(%H-%M-%S.%f)")
os.mkdir('newWorkbookFile_%s' % timestampStr)
newworkbookFileLocation = location + '\\' + 'newWorkbookFile_%s' % timestampStr
newworkbookFile = newworkbookFileLocation + '\\' + "%s"  % name
shutil.copy(workbookFile, newworkbookFileLocation)
renamednewworkbookFile = os.path.splitext(newworkbookFile)[0] + "_%s" % (timestampStr + os.path.splitext(newworkbookFile)[1])
os.rename(newworkbookFile, renamednewworkbookFile)
workbookFile = renamednewworkbookFile

print("\n\n---------------------------------------------------------------------------------")
print("Inputting the initial values")
print("---------------------------------------------------------------------------------")
while True:
    stageDirectory = input("Enter the stage directory: ")
    if os.path.isdir(stageDirectory):
        break
    else:
        print("!", '-' * 77, "!")
        print("Directory does not exist. Please try again.")
        print("!", '-' * 77, "!\n")

# ---------------------------------------------------------------------------------
# Keep track of all Folders present
# ---------------------------------------------------------------------------------
folderQueue = [i for i in os.listdir(
    stageDirectory) if os.path.isdir(stageDirectory + '\\' + i)]

# ---------------------------------------------------------------------------------
# Traverse through stage directory and each sub-folder to get to the out files
# ---------------------------------------------------------------------------------
outFiles = []
currentOutFile = [i for i in os.listdir(stageDirectory) if i.endswith('.out')]
for files in currentOutFile:
    outFiles.append(files)

while len(folderQueue) > 0:
    currentDB = folderQueue.pop(0)
    currentOutFile = [i for i in os.listdir(
        stageDirectory + '\\' + currentDB) if i.endswith('.out')]
    for files in currentOutFile:
        outFiles.append(currentDB + '\\' + files)

while True:
    environmentName = input("Enter the Environment name (TEST/PROD/DEV): ").upper()
    if environmentName not in ('TEST', 'PROD', 'DEV'):
        print("!", '-' * 77, "!")
        print("Incorrect environment. Please try again.")
        print("!", '-' * 77, "!\n")
    else:
        break
plannedDataGrowth = input("Enter the Planned Data Growth Value % [10]: ")
if len(plannedDataGrowth)==0:
    plannedDataGrowth = str(10)
plannedCPUGrowth = input("Enter the Planned CPU Growth Value % [10]: ")
if len(plannedCPUGrowth)==0:
    plannedCPUGrowth = str(10)
print("\n" * 2)

# ---------------------------------------------------------------------------------
# Read all the server model names from servers_m-values.csv
# ---------------------------------------------------------------------------------
stage_directory = stageDirectory
outFilesString = convertToString(outFiles)
IOGrowth = plannedCPUGrowth
outFiles = list(outFilesString.split(" "))
add_to_excel = {} # Dict of all servers from all AWR files. Key - ServerName, value - ServerModelName
server_list = []  # List of all servers, scope - each AWR file
mean_server = {} # Dict of all servers (keys) containing list of all os_cpu_max (values)
allservers = []  # List of all servers
sum_memUsed = {} # List of all servers with their memUsed
row = [[]]
filename = os.path.join(os.path.dirname(__file__), "servers_m-values.csv")
# D:\Work\Automation\Sizing\servers_m-values.csv

# ---------------------------------------------------------------------------------
# Get all ServerModelNames from servers_m-values.csv file
# ---------------------------------------------------------------------------------
modelNames = []
with open(filename) as modelfile:
    reader = csv.reader(modelfile)
    for line in reader:
        # print(line[1])
        modelNames.append(line[1])
# print(modelNames)

# ---------------------------------------------------------------------------------
# Function to store required values to variable to append to workbook file
# ---------------------------------------------------------------------------------
def insertToCSV(environmentName, server, serverModelName, physicalMemory):
    row.append([environmentName, server, serverModelName, physicalMemory])

# ---------------------------------------------------------------------------------
# Function to input Server Model Name if not entered for particular host
# ---------------------------------------------------------------------------------
def getServerModelNames(initialServerName):
    res = [i for i in modelNames if initialServerName in i]
    if len(res)==0:
        return 0
    print("\nThe list of Server Model Names based on the keyword are:\n")
    print("!", '-' * 77, "!")
    for i in res:
        print(i)
    print("!", '-' * 77, "!")
    while True:
        finalServerName = input("\nEnter the Server Model Name fetched from the above list:\n")
        if finalServerName not in modelNames:
            print("!", '-' * 77, "!")
            print("Incorrect Model Server Name. Please try again.")
            print("!", '-' * 77, "!")
        else:
            return finalServerName
            break


def getAllValues(f1):
    for line in f1:
        if re.search(r'^DB_NAME*', line):
            databaseName = line
            databaseName = databaseName[7:]
            databaseName = databaseName.strip()
        if re.search(r'^HOSTS*', line):
            server = line
            server = server[5:]
            server = server.strip()
            server_list = server.split(',')
        if re.search(r'^INSTANCES*', line):
            instances = line
            instances = instances[9:]
            instances = int(instances.strip())
        if re.search(r'^PHYSICAL_MEMORY_GB*', line):
            physicalMemory = line
            physicalMemory = physicalMemory[60:]
            physicalMemory = physicalMemory.replace(',', '.')
            physicalMemory = int(normal_round(float(physicalMemory.strip())))
    
    for servers in server_list:
        if servers not in add_to_excel:
            add_to_excel[servers] = None
            mean_server[servers] = []
            sum_memUsed[servers] = []
    
    # ---------------------------------------------------------------------------------
    # Get BEGIN-MAIN-METRICS & END-MAIN-METRICS line number
    # ---------------------------------------------------------------------------------
    begin_main_metrics = 0
    end_main_metrics = 0
    for num, line in enumerate(f1, 1):
        if '~~BEGIN-MAIN-METRICS~~' in line:
            begin_main_metrics = num + 3
            check_string = num + 1 # Line where string exists
        if '~~END-MAIN-METRICS~~' in line:
            end_main_metrics = num-2
            break
    start_os_cpu_max = f1[check_string].find("os_cpu_max") + 1 # Find the starting position of string "os_cpu_max"
    end_os_cpu_max = start_os_cpu_max + len("os_cpu_max") - 1 # Find the ending position of string "os_cpu_max"
    
    start_inst = f1[check_string].find("inst") + 1 # Find the starting position of string "inst"
    end_inst = start_inst + 9 # Find the ending position of string "inst"
    
    print("\n\n---------------------------------------------------------------------------------")
    print("Database %s from file %s" % (databaseName, outFile))
    print("Fetching CPU and Memory Utilization Percentage ...")
    os_cpu_max_list = []  # List of all os_cpu_max entries from an AWR file
    avg_os_cpu_max = 0 # Variable to hold the average of all os_cpu_max from an AWR file
    # Iterate through the above 2 lines
    for i in range(begin_main_metrics, end_main_metrics):
        # Gets the entire line
        line = f1[i]
        # Slices the line to include only os_cpu_max
        os_cpu_max = line[start_os_cpu_max:end_os_cpu_max]
        os_cpu_max = os_cpu_max.replace(',', '.')
        # Removes extra spaces
        os_cpu_max = os_cpu_max.strip()
        # Convert from string to float
        os_cpu_max = float(os_cpu_max)
        # Add os_cpu_max to list
        os_cpu_max_list.append(os_cpu_max)

    print("Finding average of os_cpu_max from above list ...")
    avg_os_cpu_max = statistics.mean(os_cpu_max_list)

    # Append the average os_cpu_max of the AWR file as value to the corresponding key (ServerName)
    for servers in server_list:
        mean_server[servers].append(avg_os_cpu_max)
    
    # time.sleep(3)

    # ---------------------------------------------------------------------------------
    # Get BEGIN-MEMORY and END-MEMORY line number
    # ---------------------------------------------------------------------------------
    begin_memory = 0
    end_memory = 0
    check_string = 0
    for num, line in enumerate(f1, 1):
        if '~~BEGIN-MEMORY~~' in line:
            begin_memory = num + 3
            check_string = num + 1 # Line where string exists
        if '~~END-MEMORY~~' in line:
            end_memory = num-2
            break
    
    start_sga = f1[check_string].find("SGA") - 3 # Find the starting position of string "SGA"
    end_sga = start_sga + len("SGA") + 3 # Find the ending position of string "SGA"

    start_pga = f1[check_string].find("PGA") - 3 # Find the starting position of string "PGA"
    end_pga = start_pga + len("PGA") + 3 # Find the ending position of string "PGA"

    start_snap = f1[check_string].find("SNAP_ID") # Find the starting position of string "snap"
    end_snap = start_snap + 7 # Find the ending position of string "snap"

    print("Fetching SGA and PGA ...")
    print("Calculating Memory Utilization Percentage ...")
    print("---------------------------------------------------------------------------------")
    sga_list = {}
    pga_list = {}
    max_sga = 0
    max_pga = 0
    # Iterate through the above 2 lines
    for i in range(begin_memory, end_memory):
        # Gets the entire line
        line = f1[i]
        # Slices the line to include only SGAs & PGAs. Replace ',' with '.' (if file is corrupted)
        sga = line[start_sga:end_sga]
        sga = sga.replace(',', '.')
        pga = line[start_pga:end_pga]
        pga = pga.replace(',', '.')
        # Removes extra spaces
        sga = sga.strip()
        pga = pga.strip()
        # Convert from string to float
        sga = float(assignZero(sga))
        pga = float(assignZero(pga))

        # Do above for snap, add to dictionary as key and give a initial value of 0
        snap = line[start_snap:end_snap]
        snap = snap.strip()
        snap = int(snap)
        if snap not in sga_list:
            sga_list[snap] = 0
        if snap not in pga_list:
            pga_list[snap] = 0

        # Add SGA and PGA to list for the particular snap ID
        sga_list[snap] += sga
        pga_list[snap] += pga


    # Get maximum of SGA and PGA grouped by snap
    keymax_sga = max(sga_list.keys(), key=(lambda k: sga_list[k]))
    max_sga = sga_list[keymax_sga]
    keymax_pga = max(pga_list.keys(), key=(lambda k: pga_list[k]))
    max_pga = pga_list[keymax_pga]

    # Apply calculation for total memUsed
    memUsed = max_sga + max_pga  

    # Divide by number of instances
    memUsed = memUsed/instances

    # Apply calculation for percentage
    memUsedPercentage = normal_round((memUsed/physicalMemory) * 100)

    # Append the memUsed to the corresponding key (ServerName)
    for servers in server_list:
        sum_memUsed[servers].append(memUsedPercentage)
    # time.sleep(3)

    for i in range(instances):
        print("\n\n---------------------------------------------------------------------------------")
        print("Database: %s, Host: %s" % (databaseName, server_list[i]))
        print("---------------------------------------------------------------------------------")
        if add_to_excel[server_list[i]] == None:
            while True:
                initialServerName = input("Enter the Server Model Name keyword (Press 'Enter' to see all): ").upper()
                serverModelName = getServerModelNames(initialServerName)
                if serverModelName==0:
                    print("!", '-' * 77, "!")
                    print("Keyword fetched an empty list. Please try again.")
                    print("!", '-' * 77, "!\n")
                else:
                    add_to_excel[server_list[i]] = serverModelName
                    break
        else:
            print("Fetching ServerModelName from file ...")
            serverModelName = add_to_excel[server_list[i]]
        if server_list[i] not in allservers:
            allservers.append(server_list[i])
            insertToCSV(environmentName, server_list[i], serverModelName, physicalMemory)


# ---------------------------------------------------------------------------------
# Function to upload the workbook
# ---------------------------------------------------------------------------------
def uploadMappingSheet():
    try:
        xlApp = win32com.client.DispatchEx('Excel.Application')
        xlsPath = os.path.expanduser(
            workbookFile)
        wb = xlApp.Workbooks.Open(Filename=xlsPath)
        xlApp.Run('Upload')
        wb.Save()
        xlApp.Quit()
        print("Macro ran successfully!")
        print("\nIn case of any errors while uploading, please check file %s for mandatory fields." % workbookFile)
        print("Upload Manually once done: RAS -> Upload")
    except Exception as e:
        print(e)
        print("Error found while running the excel macro!")
        xlApp.Quit()


# ---------------------------------------------------------------------------------
# Main Function
# ---------------------------------------------------------------------------------
for outFile in outFiles:
    file = open(stage_directory + '\\' + outFile)
    f1 = file.readlines()
    getAllValues(f1)
    file.close()

# ---------------------------------------------------------------------------------
# Append to Workbook file
# ---------------------------------------------------------------------------------
wbk = openpyxl.load_workbook(
    filename=workbookFile, read_only=False, keep_vba=True)
wks = wbk['Database Servers']
for r in range(1, len(row)):
    x = r + 1
    wks.cell(row=x, column=1).value = row[r][0]
    wks.cell(row=x, column=2).value = row[r][1]
    wks.cell(row=x, column=3).value = row[r][2]
    wks.cell(row=x, column=6).value = int(normal_round(
        max(mean_server[row[r][1]])))
    wks.cell(row=x, column=7).value = row[r][3]
    wks.cell(row=x, column=8).value = int(
        sum(sum_memUsed[row[r][1]]))
    wks.cell(row=x, column=11).value = int(plannedCPUGrowth)
wbk.save(workbookFile)
wbk.close

# ---------------------------------------------------------------------------------
# Thread 1 : call the python script for Reading Database inputs from out files
# ---------------------------------------------------------------------------------
subprocess.call(['python', 'ReadDatabaseInputs.py', stageDirectory, convertToString(outFiles), environmentName, plannedDataGrowth, workbookFile])

# ---------------------------------------------------------------------------------
# Thread 2: call the python script for Reading Server inputs from out files
# ---------------------------------------------------------------------------------
subprocess.call(['python', 'ReadServerInputs.py', stageDirectory, convertToString(outFiles), environmentName, plannedCPUGrowth, workbookFile])


print("\n\n---------------------------------------------------------------------------------")
print("Uploading workbook ...")
uploadMappingSheet()
print("---------------------------------------------------------------------------------")

print("\n\n---------------------------------------------------------------------------------")
print("WrapperScript.py SCRIPT END!")
print("---------------------------------------------------------------------------------")


# ---------------------------------------------------------------------------------
# Print time taken for the entire execution
# ---------------------------------------------------------------------------------
endTime = datetime.datetime.now()
timeDifference = (date_diff_in_Minutes(endTime, startTime))
print("\n\nExecution Time: %.2f minutes" % timeDifference)